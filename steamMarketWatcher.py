import time
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import getpass
from os import path
import os
import datetime
desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')


	
currency = 3
#Currency: 1 -> USD
#Currency 3 -> EURO

cases = ["Operation Breakout Weapon Case", "Chroma 3 Case"]
itemIDs = ["14962905","149865785"]

def getCount(itemID):
	r = requests.get("https://steamcommunity.com/market/itemordershistogram?country=DE&language=german&currency=3&item_nameid=" + itemID + "&two_factor=0")
	soup = BeautifulSoup(r.content,features="lxml")

	mydivs = soup.findAll("span")
	count = mydivs[0].text.split(" ")[0]
	return count

def getPrice(itemID,currency):
	r = requests.get("https://steamcommunity.com/market/itemordershistogram?country=DE&language=german&currency="+str(currency)+"&item_nameid=" + itemID + "&two_factor=0")
	soup = BeautifulSoup(r.content,features="lxml")
	mydivs = soup.findAll("span")
	price = mydivs[0].text.split(" ")[3].split("\\")[0]

	return price

wb = openpyxl.load_workbook(desktop + "\Steam Market Watcher\excelFiles\marketWatcher.xlsx")
for i, sheet in enumerate(wb):
	max_rows = str(sheet.max_row+1)
	sheet["A"+max_rows] = getCount(itemIDs[i])
	sheet["B"+max_rows] = getPrice(itemIDs[i],currency)
	sheet["C"+max_rows] = time.time()
	sheet["D"+max_rows] = str(datetime.datetime.now())[:16]

wb.save(desktop + "\Steam Market Watcher\excelFiles\marketWatcher.xlsx")
