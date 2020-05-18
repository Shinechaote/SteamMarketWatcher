import time
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import datetime
from os import path
import os

desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')


	
currency = 3
#Currency: 1 -> USD
#Currency 3 -> EURO

cases = ["Operation Breakout Weapon Case", "Chroma 3 Case"]
itemIDs = ["14962905","149865785"]

def getCount(itemID):
	r = requests.get("https://steamcommunity.com/market/itemordershistogram?country=DE&language=german&currency=3&item_nameid=" + itemID + "&two_factor=0")
	soup = BeautifulSoup(r.content,features="lxml")

	mydivs = soup.findAll("span")
	count = mydivs[0].text.split(" ")[0]
	return count

def getPrice(itemID,currency):
	r = requests.get("https://steamcommunity.com/market/itemordershistogram?country=DE&language=german&currency="+str(currency)+"&item_nameid=" + itemID + "&two_factor=0")
	soup = BeautifulSoup(r.content,features="lxml")
	mydivs = soup.findAll("span")
	price = mydivs[0].text.split(" ")[3].split("\\")[0]

	return price

end = True
start_time = int(time.time())-600
wb = Workbook()
for i in range(len(cases)):
	ws = wb.create_sheet(cases[i])
	ws.title = cases[i]
if(len(wb.sheetnames) > len(cases)):
	wb.remove(wb["Sheet"])
while(end):
	if(int(time.time())-start_time > 600):
		print(str((int(time.time())%(24*3600)//3600)+2)+":"+str((int(time.time())%(3600)//60)))
		for i, sheet in enumerate(wb):
			max_rows = str(sheet.max_row+1)
			if(sheet.max_row == 1):
				max_rows = str(int(max_rows)-1)
			price = getPrice(itemIDs[i],currency)
			count = getCount(itemIDs[i])
			print(cases[i]+":")
			print("Count: " + str(count))
			print("Price: " + str(price))
			print()
			print()
			sheet["A"+max_rows] = count
			sheet["B"+max_rows] = price
			sheet["C"+max_rows] = time.time()%(24*3600)
			sheet["D"+max_rows] = str(datetime.datetime.now())[:16]
		start_time = time.time()
		wb.save(desktop + "\Steam Market Watcher\excelFiles\marketWatcher" + str(datetime.datetime.today())[:10] + ".xlsx")

	elif(int(time.time())%(24*3600)>(22*3600)):
		end = False

wb.save(desktop + "\Steam Market Watcher\excelFiles\marketWatcher" + str(datetime.datetime.today())[:10] + ".xlsx")